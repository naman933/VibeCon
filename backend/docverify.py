import pdfplumber
import re
import csv
import io
import openpyxl
from concurrent.futures import ThreadPoolExecutor
from difflib import SequenceMatcher
from datetime import datetime

COLUMN_ALIASES = {
    'application_id': ['application id', 'app id', 'application_id', 'id', 'application no', 'app no'],
    'name': ['candidate name', 'name', 'student name', 'applicant name', 'full name'],
    'cat_reg_no': ['cat registration number', 'cat reg no', 'cat registration no', 'cat id',
                   'registration number', 'reg no', 'registration no', 'cat reg'],
    'date_of_test': ['date of test', 'test date', 'date', 'exam date'],
    'varc_percentile': ['varc percentile', 'varc', 'verbal ability', 'varc %ile',
                        'varc percentile score', 'verbal ability & reading comprehension'],
    'dilr_percentile': ['dilr percentile', 'dilr', 'data interpretation', 'dilr %ile',
                        'dilr percentile score', 'data interpretation & logical reasoning'],
    'qa_percentile': ['qa percentile', 'qa', 'quantitative ability', 'qa %ile',
                      'qa percentile score', 'quantitative aptitude'],
    'overall_percentile': ['overall percentile', 'overall', 'total percentile', 'overall %ile',
                           'aggregate percentile', 'total'],
}


def normalize_column(col_name):
    col_lower = col_name.strip().lower()
    for field, aliases in COLUMN_ALIASES.items():
        if col_lower in aliases:
            return field
    return None


def parse_number(val):
    if val is None:
        return None
    s = str(val).strip()
    try:
        return float(s)
    except ValueError:
        nums = re.findall(r'\d+\.?\d*', s)
        return float(nums[-1]) if nums else None


def normalize_date(date_str):
    if not date_str:
        return None
    s = str(date_str).strip()
    # Remove ordinal suffixes (1st, 2nd, 3rd, 4th, etc.)
    s_clean = re.sub(r'(\d+)(st|nd|rd|th)\b', r'\1', s, flags=re.IGNORECASE)
    for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%B %d, %Y', '%b %d, %Y',
                '%d %B %Y', '%d %b %Y', '%A, %B %d, %Y', '%A, %d %B %Y',
                '%d %b %Y', '%b %d %Y']:
        for candidate_str in [s_clean, s]:
            cleaned = re.sub(r'(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),?\s*', '', candidate_str, flags=re.IGNORECASE).strip()
            try:
                return datetime.strptime(cleaned, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return s


def normalize_numeric_string(val):
    """Strip trailing .0 from numeric strings (e.g. '21161013.0' -> '21161013')"""
    s = str(val).strip()
    if re.match(r'^\d+\.0$', s):
        return s[:-2]
    # Handle scientific notation like 2.1161013E7
    try:
        f = float(s)
        if f == int(f) and ('e' in s.lower() or 'E' in s):
            return str(int(f))
    except (ValueError, OverflowError):
        pass
    return s


def parse_formdata(file_bytes, filename):
    candidates = []
    rows = []

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        headers_raw = [str(cell.value or '').strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers_raw, [str(v) if v is not None else '' for v in row])))
        wb.close()
    else:
        text = file_bytes.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)

    for row in rows:
        candidate = {}
        for raw_col, raw_val in row.items():
            field = normalize_column(raw_col)
            if field:
                candidate[field] = raw_val.strip() if raw_val else ''
        if candidate.get('name') or candidate.get('cat_reg_no'):
            # Normalize numeric fields
            if candidate.get('cat_reg_no'):
                candidate['cat_reg_no'] = normalize_numeric_string(candidate['cat_reg_no'])
            if candidate.get('application_id'):
                candidate['application_id'] = normalize_numeric_string(candidate.get('application_id', ''))
            for pf in ['varc_percentile', 'dilr_percentile', 'qa_percentile', 'overall_percentile']:
                if pf in candidate:
                    candidate[pf] = parse_number(candidate[pf])
            candidate['date_of_test'] = normalize_date(candidate.get('date_of_test', ''))
            if 'application_id' not in candidate:
                candidate['application_id'] = ''
            candidates.append(candidate)

    return candidates


def parse_single_pdf(pdf_bytes, filename):
    result = {
        'filename': filename,
        'cat_reg_no': None,
        'candidate_name': None,
        'date_of_test': None,
        'varc_percentile': None,
        'dilr_percentile': None,
        'qa_percentile': None,
        'overall_percentile': None,
        'confidence': 'high',
        'error': None,
    }

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            all_text = ''
            all_tables = []
            for page in pdf.pages:
                text = page.extract_text() or ''
                all_text += text + '\n'
                tables = page.extract_tables()
                if tables:
                    all_tables.extend(tables)

        if len(all_text.strip()) < 20:
            result['confidence'] = 'low'
            result['error'] = 'Insufficient text extracted - may be image-based PDF'
            return result

        # Extract CAT Registration Number
        reg_patterns = [
            r'(?:CAT\s*)?Registration\s*(?:Number|No\.?)\s*[:=\-]?\s*([A-Z0-9][-A-Z0-9]+)',
            r'Reg\.?\s*(?:Number|No\.?)\s*[:=\-]?\s*([A-Z0-9][-A-Z0-9]+)',
            r'([A-Z]\d{2}[-]\d{5,})',
        ]
        for pat in reg_patterns:
            m = re.search(pat, all_text, re.IGNORECASE)
            if m:
                result['cat_reg_no'] = m.group(1).strip()
                break

        # Extract candidate name
        name_patterns = [
            r'Name\s*of\s*(?:the\s*)?Candidate\s*[:=\-]?\s*([A-Z][A-Za-z\s.]+?)(?:\n|$)',
            r'(?:Candidate|Student|Applicant)\s*(?:Name|\'s Name)\s*[:=\-]?\s*([A-Z][A-Za-z\s.]+?)(?:\n|$)',
            r'Name\s*[:=\-]?\s*([A-Z][A-Za-z\s.]{3,40}?)(?:\n|$)',
        ]
        for pat in name_patterns:
            m = re.search(pat, all_text)
            if m:
                result['candidate_name'] = m.group(1).strip()
                break

        # Extract Date of Test
        date_patterns = [
            r'Date\s*of\s*(?:the\s*)?Test\s*[:=\-]?\s*(.+?)(?:\n|$)',
            r'Test\s*Date\s*[:=\-]?\s*(.+?)(?:\n|$)',
            r'Exam\s*Date\s*[:=\-]?\s*(.+?)(?:\n|$)',
        ]
        for pat in date_patterns:
            m = re.search(pat, all_text, re.IGNORECASE)
            if m:
                result['date_of_test'] = normalize_date(m.group(1).strip())
                break

        # Extract percentiles from tables
        # CAT scorecards typically have a table structure like:
        # Row 0: Section headers (Verbal Ability, Data Interpretation, Quantitative, Total)
        # Row 1: Sub-headers (may repeat section names)
        # Row 2: Column labels (Scaled Score, Percentile, ...)
        # Row 3: Values (31.75, 93.21, 25.79, 96.20, 15.08, 86.70, 72.62, 95.10)

        def normalize_cell(cell):
            """Remove extra spaces within words (OCR artifact: 'Percent ile' -> 'Percentile')"""
            return re.sub(r'\s+', ' ', str(cell or '')).strip().lower()

        def contains_percentile(cell_str):
            """Check if cell contains 'percentile' even with OCR spaces"""
            cleaned = re.sub(r'\s+', '', cell_str.lower())
            return 'percentile' in cleaned and 'scaledscore' not in cleaned

        section_map = {
            'varc': 'varc_percentile',
            'verbal': 'varc_percentile',
            'reading comprehension': 'varc_percentile',
            'dilr': 'dilr_percentile',
            'data interpretation': 'dilr_percentile',
            'logical reasoning': 'dilr_percentile',
            'qa': 'qa_percentile',
            'quantitative': 'qa_percentile',
            'overall': 'overall_percentile',
            'aggregate': 'overall_percentile',
            'total': 'overall_percentile',
        }

        def match_section(cell_str):
            """Match cell text to a section field, handling OCR spaces"""
            cleaned = re.sub(r'\s+', '', cell_str).lower()
            for keyword, field in section_map.items():
                kw_clean = keyword.replace(' ', '')
                if kw_clean in cleaned:
                    return field
            return None

        for table in all_tables:
            if len(table) < 2:
                continue

            # Skip layout tables (single-cell rows with long text = containers, not data)
            max_cols = max(len(row) for row in table) if table else 0
            if max_cols < 2:
                continue

            # Strategy 1: Columnar layout — map section header columns to percentile values
            percentile_col_indices = []
            label_row_idx = None
            for ri, row in enumerate(table):
                for ci, cell in enumerate(row):
                    cell_str = normalize_cell(cell)
                    if contains_percentile(cell_str):
                        percentile_col_indices.append(ci)
                        label_row_idx = ri

            if percentile_col_indices and label_row_idx is not None:
                # Find the data row (row after label row with numeric values)
                data_row = None
                for ri in range(label_row_idx + 1, len(table)):
                    nums = [parse_number(c) for c in table[ri] if parse_number(c) is not None]
                    if len(nums) >= 3:
                        data_row = table[ri]
                        break

                if data_row:
                    # Map section headers from earlier rows to column positions
                    col_to_section = {}
                    for ri in range(label_row_idx):
                        row = table[ri]
                        for ci, cell in enumerate(row):
                            cell_str = normalize_cell(cell)
                            if not cell_str:
                                continue
                            field = match_section(cell_str)
                            if field:
                                col_to_section[ci] = field

                    # Also check the label row itself for "Overall" in "Overall Percentile"
                    for ci, cell in enumerate(table[label_row_idx]):
                        cell_str = normalize_cell(cell)
                        field = match_section(cell_str)
                        if field and ci not in col_to_section:
                            col_to_section[ci] = field

                    # Extract percentile values from data row using column mapping
                    for pci in sorted(percentile_col_indices):
                        if pci < len(data_row):
                            val = parse_number(data_row[pci])
                            if val is not None and 0 <= val <= 100:
                                # Find which section this column belongs to
                                matched_field = None
                                # Check if this column itself has a section mapping
                                if pci in col_to_section:
                                    matched_field = col_to_section[pci]
                                else:
                                    # Find nearest section to the left
                                    for ci in sorted(col_to_section.keys(), reverse=True):
                                        if ci <= pci:
                                            matched_field = col_to_section[ci]
                                            break
                                if matched_field and result[matched_field] is None:
                                    result[matched_field] = val

                    continue  # Skip strategy 2 for this table

            # Strategy 2: Row-based layout (original logic)
            header_row = table[0] if table else []
            percentile_col_idx = None
            if header_row:
                for idx, cell in enumerate(header_row):
                    if cell and contains_percentile(normalize_cell(cell)):
                        percentile_col_idx = idx
                        break

            for row in table:
                if not row:
                    continue
                row_text = ' '.join(str(cell) for cell in row if cell).lower()
                for keyword, field in section_map.items():
                    if keyword in row_text:
                        if percentile_col_idx is not None and percentile_col_idx < len(row):
                            val = parse_number(row[percentile_col_idx])
                            if val is not None and 0 <= val <= 100:
                                result[field] = val
                        else:
                            numbers = [parse_number(cell) for cell in row if parse_number(cell) is not None]
                            valid = [n for n in numbers if 0 <= n <= 100]
                            if valid:
                                result[field] = valid[-1]
                        break

        # Fallback: regex on raw text
        for keyword, field in [('varc|verbal ability', 'varc_percentile'),
                               ('dilr|data interpretation', 'dilr_percentile'),
                               ('qa|quantitative', 'qa_percentile'),
                               ('overall|aggregate', 'overall_percentile')]:
            if result[field] is None:
                pattern = rf'(?:{keyword})[^\n]*?(\d{{1,3}}\.?\d*)'
                matches = re.findall(pattern, all_text, re.IGNORECASE)
                if matches:
                    for m_val in reversed(matches):
                        v = float(m_val)
                        if 0 <= v <= 100:
                            result[field] = v
                            break

        # Assess confidence
        extracted = sum(1 for v in [result['cat_reg_no'], result['varc_percentile'],
                                    result['dilr_percentile'], result['qa_percentile'],
                                    result['overall_percentile']] if v is not None)
        if extracted <= 2:
            result['confidence'] = 'low'
        elif extracted <= 3:
            result['confidence'] = 'medium'

    except Exception as e:
        result['confidence'] = 'error'
        result['error'] = str(e)

    return result


def parse_pdfs_parallel(pdf_list, max_workers=8):
    scorecards = []
    with ThreadPoolExecutor(max_workers=min(max_workers, len(pdf_list) or 1)) as executor:
        futures = {executor.submit(parse_single_pdf, p['bytes'], p['filename']): p['filename']
                   for p in pdf_list}
        for future in futures:
            scorecards.append(future.result())
    return scorecards


def name_similarity(a, b):
    if not a or not b:
        return 0
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def match_and_verify(candidates, scorecards):
    results = []
    matched_pdfs = set()

    for cand in candidates:
        cand_reg = normalize_numeric_string((cand.get('cat_reg_no') or '').strip()).upper()
        best_match = None
        match_method = None

        # Primary: match by CAT Registration Number
        for sc in scorecards:
            sc_reg = normalize_numeric_string((sc.get('cat_reg_no') or '').strip()).upper()
            if cand_reg and sc_reg and cand_reg == sc_reg:
                best_match = sc
                match_method = 'cat_reg_no'
                break

        # Fallback: match by name
        if not best_match:
            best_score = 0
            for sc in scorecards:
                if sc['filename'] in matched_pdfs:
                    continue
                sim = name_similarity(cand.get('name', ''), sc.get('candidate_name', ''))
                if sim > 0.8 and sim > best_score:
                    best_score = sim
                    best_match = sc
                    match_method = 'name_fuzzy'

        entry = {
            'application_id': cand.get('application_id', ''),
            'name': cand.get('name', ''),
            'cat_reg_no_form': cand.get('cat_reg_no', ''),
            'match_method': match_method,
            'pdf_filename': best_match['filename'] if best_match else None,
            'pdf_confidence': best_match['confidence'] if best_match else None,
            'status': 'VERIFIED',
            'fields': {},
            'form_values': {},
            'pdf_values': {},
            'issues': [],
        }

        if not best_match:
            entry['status'] = 'MISSING_DOCUMENT'
            entry['issues'].append('No matching PDF scorecard found')
            results.append(entry)
            continue

        matched_pdfs.add(best_match['filename'])

        if best_match['confidence'] == 'error':
            entry['status'] = 'NEEDS_REVIEW'
            entry['issues'].append(f"PDF read error: {best_match.get('error', 'Unknown')}")
            results.append(entry)
            continue

        if best_match['confidence'] == 'low':
            entry['status'] = 'NEEDS_REVIEW'
            entry['issues'].append('Low OCR confidence on PDF')

        # Compare fields
        def compare_exact(field_label, form_val, pdf_val):
            f = str(form_val or '').strip().upper()
            p = str(pdf_val or '').strip().upper()
            entry['form_values'][field_label] = str(form_val or '') if form_val else None
            entry['pdf_values'][field_label] = str(pdf_val or '') if pdf_val else None
            if not f and not p:
                entry['fields'][field_label] = 'both_empty'
            elif not p:
                entry['fields'][field_label] = 'not_in_pdf'
                entry['issues'].append(f'{field_label} not extracted from PDF')
            elif not f:
                entry['fields'][field_label] = 'not_in_form'
            elif f == p:
                entry['fields'][field_label] = 'match'
            else:
                entry['fields'][field_label] = 'mismatch'
                entry['issues'].append(f'{field_label} mismatch: Form="{form_val}" vs PDF="{pdf_val}"')

        def compare_percentile(field_label, form_val, pdf_val, tolerance=0.1):
            fv = parse_number(form_val)
            pv = parse_number(pdf_val)
            entry['form_values'][field_label] = fv
            entry['pdf_values'][field_label] = pv
            if fv is None and pv is None:
                entry['fields'][field_label] = 'both_empty'
            elif pv is None:
                entry['fields'][field_label] = 'not_in_pdf'
                entry['issues'].append(f'{field_label} not extracted from PDF')
            elif fv is None:
                entry['fields'][field_label] = 'not_in_form'
            elif abs(fv - pv) <= tolerance:
                entry['fields'][field_label] = 'match'
            else:
                entry['fields'][field_label] = 'mismatch'
                entry['issues'].append(f'{field_label} mismatch: Form={fv} vs PDF={pv}')

        compare_exact('cat_reg_no', cand.get('cat_reg_no'), best_match.get('cat_reg_no'))
        compare_exact('date_of_test', cand.get('date_of_test'), best_match.get('date_of_test'))
        compare_percentile('varc', cand.get('varc_percentile'), best_match.get('varc_percentile'))
        compare_percentile('dilr', cand.get('dilr_percentile'), best_match.get('dilr_percentile'))
        compare_percentile('qa', cand.get('qa_percentile'), best_match.get('qa_percentile'))
        compare_percentile('overall', cand.get('overall_percentile'), best_match.get('overall_percentile'))

        mismatches = sum(1 for v in entry['fields'].values() if v == 'mismatch')
        not_extracted = sum(1 for v in entry['fields'].values() if v == 'not_in_pdf')

        if mismatches > 0:
            entry['status'] = 'DISCREPANCY'
        elif not_extracted > 2 or best_match['confidence'] == 'low':
            entry['status'] = 'NEEDS_REVIEW'
        elif not_extracted > 0:
            entry['status'] = 'NEEDS_REVIEW'
        else:
            entry['status'] = 'VERIFIED'

        results.append(entry)

    # Check for unmatched PDFs
    for sc in scorecards:
        if sc['filename'] not in matched_pdfs:
            results.append({
                'application_id': '',
                'name': sc.get('candidate_name', 'Unknown'),
                'cat_reg_no_form': '',
                'match_method': None,
                'pdf_filename': sc['filename'],
                'pdf_confidence': sc['confidence'],
                'status': 'NEEDS_REVIEW',
                'fields': {},
                'issues': ['PDF not matched to any candidate in Form A'],
            })

    # Check duplicates
    reg_counts = {}
    for r in results:
        reg = r.get('cat_reg_no_form', '')
        if reg:
            reg_counts[reg] = reg_counts.get(reg, 0) + 1
    for r in results:
        reg = r.get('cat_reg_no_form', '')
        if reg and reg_counts.get(reg, 0) > 1:
            r['status'] = 'NEEDS_REVIEW'
            r['issues'].append('Duplicate CAT Registration Number detected')

    # Summary
    summary = {
        'total': len(results),
        'verified': sum(1 for r in results if r['status'] == 'VERIFIED'),
        'needs_review': sum(1 for r in results if r['status'] == 'NEEDS_REVIEW'),
        'discrepancy': sum(1 for r in results if r['status'] == 'DISCREPANCY'),
        'missing_document': sum(1 for r in results if r['status'] == 'MISSING_DOCUMENT'),
    }

    return {'summary': summary, 'results': results}
